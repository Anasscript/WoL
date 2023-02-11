import os
import re
import subprocess
try:
    from termcolor import colored
except ImportError:
    subprocess.run(["pip", "install", "termcolor"])
    from termcolor import colored
    import openpyxl
except ImportError:
    subprocess.run(["pip", "install", "openpyxl"])
    import openpyxl
try:
    import art
except ImportError:
    subprocess.run(["pip", "install", "art"])
    import art
    from art import *
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from art import *
import time
import os
import sys
try:
    from tqdm import tqdm
except ImportError:
    subprocess.run(["pip", "install", "tqdm"])
    from tqdm import tqdm
from tqdm import tqdm

print("\nCaricamento completato")

def textart(scritta, mod):
    if mod==1:
        tprint(scritta, font="rand-small")
    if mod==2:
        tprint(scritta, font="rand-medium")
    if mod==3:
        tprint(scritta, font="rand-large")
    else :
        tprint(scritta, font=mod)

def text_art_menu():
        textart("Turn On Manager", "georgia11")
        print("Powered for")
        textart("Piazza HR", "double")
        print("Caricamento in corso: ", end="")
        return 0

text_art_menu() # stampa una text art

for i in tqdm(range(40), desc="Caricamento in corso"):
    time.sleep(0.1)

def clear():
   os.system("clear" if os.name == "posix" else "cls")# pulisci lo schermo


def get_devices_list():
    # Crea una lista dei dispositivi connessi
    output = subprocess.run(["arp", "-a"], capture_output=True, text=True)
    devices = re.findall(r"([\w.-]+)\s+\(([\d.]+)\)", output.stdout)
    return devices

def show_devices_list(devices):
    # DMostra i dispositivi connessi
    tprint("Dispositivi connessi:", )
    for i, (device, ip) in enumerate(devices):
        print(f"{i + 1}. {device} ({ip})")

def select_devices_to_wake(devices):
    # Prompt the user to select the devices to turn on
    print("Enter the number(s) of the device(s) to turn on (separated by spaces, or 'all' to select all devices):")
    selected = input().strip()
    if selected == "all":
        return devices
    else:
        selected_indices = map(int, selected.split())
        return [devices[i - 1] for i in selected_indices]

def send_wake_on_lan(devices):
    # Send the Wake-On-LAN signal to the selected devices
    for device, ip in devices:
        mac_address = get_mac_address(ip)
        if mac_address:
            os.system(f"wakeonlan {mac_address}")

def get_mac_address(ip):
    # Get the MAC address of the device with the given IP address
    output = subprocess.run(["arp", "-n", ip], capture_output=True, text=True)
    match = re.search(r"([\da-fA-F:]+)", output.stdout)
    if match:
        return match.group(1)

def schedule_wake_on_lan(devices):
    # Prompt the user to choose between the advanced or simple scheduling mode
    print("Choose scheduling mode:")
    print("1. Advanced (export to Excel)")
    print("2. Simple (set schedule for each day of the week)")
    choice = int(input().strip())
    if choice == 1:
        export_to_excel(devices)
    elif choice == 2:
        set_simple_schedule(devices)

def export_to_excel(devices):
    # Export the list of devices and their scheduled turn-on time to an Excel file
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Scheduled Devices"
    worksheet["A1"] = "Device"
    worksheet["B1"] = "Time"
    worksheet.column_dimensions[get_column_letter(1)].width = 20
    worksheet.column_dimensions[get_column_letter(2)].width = 8
    header_font = Font(bold=True)
    worksheet["A1"].font = header_font
    worksheet["B1"].font = header_font

    for i, (device, ip) in enumerate(devices):
        worksheet.cell(row=i + 2, column=1, value=device)
        worksheet.cell(row=i + 2, column=2, value="00:00")

    workbook.save("scheduled_devices.xlsx")
    print("Scheduled devices exported to 'scheduled_devices.xlsx'.")

def set_simple_schedule(devices):
    # Set a common turn-on time for each day of the week for all devices
    print("Enter turn-on time (HH:MM):")
    time = input().strip()
    print("Turn-on time set to " + time + " for all devices.")

if __name__ == "__main__":
    devices = get_devices_list()
    show_devices_list(devices)
    selected_devices = select_devices_to_wake(devices)
    send_wake_on_lan(selected_devices)
    schedule_wake_on_lan(selected_devices)